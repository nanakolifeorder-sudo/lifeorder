import json
import re
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path("source_booking_records.xlsx")
OUTPUT = Path("legacy_import_payload.json")
TZ = timezone(timedelta(hours=8))


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def iso_dt(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=TZ).isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=TZ).isoformat()
    text = clean(value)
    if not text:
        return ""
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=TZ).isoformat()
        except ValueError:
            pass
    return text


def date_str(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean(value)


def time_str(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = clean(value)
    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    return text


def nonempty_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None and clean(cell) for cell in row):
            rows.append(list(row))
    return rows


def cell(row, index):
    return row[index] if index < len(row) else None


def map_question_type(value):
    text = clean(value)
    if text in ("單選", "多選"):
        return text
    return "文字"


def split_options(value):
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,，\n]", text) if part.strip()]


def int_or(value, fallback):
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def normalize_trigger(value):
    text = clean(value)
    table = {
        "預約成功時": "booking_created",
        "會議前提醒": "meeting_before",
        "會議後跟進": "meeting_after",
        "取消預約時": "booking_cancelled",
        "取消時": "booking_cancelled",
        "發送改期連結時": "reschedule_link",
        "名單建立": "lead_created",
        "表單送出": "lead_created",
        "未預約跟進": "lead_created",
    }
    return table.get(text, text or "lead_created")


def replace_vars(value):
    text = clean(value)
    replacements = {
        "{{客戶姓名}}": "{{name}}",
        "{{姓名}}": "{{name}}",
        "{{客戶Email}}": "{{email}}",
        "{{Email}}": "{{email}}",
        "{{客戶電話}}": "{{phone}}",
        "{{電話}}": "{{phone}}",
        "{{專案名稱}}": "{{project}}",
        "{{預約時間}}": "{{time}}",
        "{{會議連結}}": "{{meetLink}}",
        "{{顧問姓名}}": "{{consultant}}",
        "{{顧問}}": "{{consultant}}",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def main():
    wb = load_workbook(SOURCE, data_only=True)
    sheets = {ws.title: ws for ws in wb.worksheets}

    projects = []
    project_by_name = {}
    for row in nonempty_rows(sheets["專案設定"])[1:]:
        code = clean(cell(row, 0)).upper()
        if not code:
            continue
        name = clean(cell(row, 1)) or code
        project_by_name[name] = code
        projects.append(
            {
                "code": code,
                "name": name,
                "status": clean(cell(row, 2)) or "啟用",
                "mainUrl": clean(cell(row, 3)),
                "fallbackUrl": clean(cell(row, 4)),
                "bookingNotice": clean(cell(row, 5)),
                "rejectValue": clean(cell(row, 6)),
            }
        )

    questions = []
    for row in nonempty_rows(sheets["問卷題庫"])[1:]:
        project = clean(cell(row, 0)).upper()
        title = clean(cell(row, 2))
        if not project or not title:
            continue
        questions.append(
            {
                "projectCode": project,
                "sortOrder": int_or(cell(row, 1), len(questions) + 1),
                "title": title,
                "type": map_question_type(cell(row, 3)),
                "options": split_options(cell(row, 4)),
                "rejectWord": clean(cell(row, 5)),
                "isRequired": True,
            }
        )

    consultants = []
    for row in nonempty_rows(sheets["顧問團隊與權限"])[1:]:
        email = clean(cell(row, 1)).lower()
        if not email:
            continue
        calendar_ids = [item.strip() for item in clean(cell(row, 3)).split(",") if item.strip()]
        consultants.append(
            {
                "name": clean(cell(row, 0)),
                "email": email,
                "password": clean(cell(row, 2)) or "1111",
                "calendarId": calendar_ids[0] if calendar_ids else "primary",
                "accepting": clean(cell(row, 4)) or "否",
                "weight": int(cell(row, 5) or 50),
                "permissions": clean(cell(row, 7)),
                "projectCodes": clean(cell(row, 8)) or "ALL",
                "meetTool": clean(cell(row, 9)) or "Google Meet",
                "timeZone": "Asia/Taipei",
                "intervalMinutes": 60,
                "bufferBeforeMinutes": 0,
                "bufferAfterMinutes": 0,
            }
        )

    weekday_map = {
        "星期日": 0,
        "週日": 0,
        "周日": 0,
        "星期一": 1,
        "週一": 1,
        "周一": 1,
        "星期二": 2,
        "週二": 2,
        "周二": 2,
        "星期三": 3,
        "週三": 3,
        "周三": 3,
        "星期四": 4,
        "週四": 4,
        "周四": 4,
        "星期五": 5,
        "週五": 5,
        "周五": 5,
        "星期六": 6,
        "週六": 6,
        "周六": 6,
    }
    availability = []
    for row in nonempty_rows(sheets["顧問排班表"])[1:]:
        name = clean(cell(row, 0))
        if not name:
            continue
        raw_date = cell(row, 1)
        day_text = clean(raw_date)
        start = time_str(cell(row, 2))
        end = time_str(cell(row, 3))
        if not start or not end:
            continue
        item = {
            "consultantName": name,
            "startTime": start,
            "endTime": end,
            "intervalMinutes": int(cell(row, 4) or 60),
            "bufferBeforeMinutes": int(cell(row, 5) or 0),
            "bufferAfterMinutes": int(cell(row, 6) or 0),
        }
        if day_text in weekday_map:
            item.update({"kind": "weekly", "dayOfWeek": weekday_map[day_text], "dateValue": ""})
        else:
            item.update({"kind": "specific", "dayOfWeek": None, "dateValue": date_str(raw_date)})
        availability.append(item)

    def project_code(project_name):
        text = clean(project_name)
        if text in project_by_name:
            return project_by_name[text]
        upper = text.upper()
        if any(project["code"] == upper for project in projects):
            return upper
        return "P01"

    def record_from(row, kind):
        meet_or_event = clean(cell(row, 9))
        event_id = ""
        meet_link = ""
        if meet_or_event.startswith("http"):
            meet_link = meet_or_event
        elif meet_or_event:
            event_id = meet_or_event
        return {
            "createdAt": iso_dt(cell(row, 0)),
            "projectName": clean(cell(row, 1)),
            "projectCode": project_code(cell(row, 1)),
            "startAt": iso_dt(cell(row, 2)),
            "clientName": clean(cell(row, 3)),
            "email": clean(cell(row, 4)).lower(),
            "phone": clean(cell(row, 5)),
            "answers": clean(cell(row, 6)),
            "consultantName": clean(cell(row, 7)),
            "status": clean(cell(row, 8)) or kind,
            "eventId": event_id,
            "meetLink": meet_link,
            "attendance": clean(cell(row, 10)),
            "dealStatus": clean(cell(row, 11)),
            "plan": clean(cell(row, 12)),
            "notes": clean(cell(row, 13)),
        }

    appointments = [
        record_from(row, "booked")
        for row in nonempty_rows(sheets["預約總紀錄"])[1:]
        if clean(cell(row, 3)) or clean(cell(row, 4))
    ]
    rejected = [
        record_from(row, "rejected")
        for row in nonempty_rows(sheets["被拒絕客戶"])[1:]
        if clean(cell(row, 3)) or clean(cell(row, 4))
    ]
    waitlist = [
        record_from(row, "waiting")
        for row in nonempty_rows(sheets["等候名單"])[1:]
        if clean(cell(row, 3)) or clean(cell(row, 4))
    ]

    email_templates = []
    for row in nonempty_rows(sheets["信件範本設定"])[1:]:
        code = clean(cell(row, 0)).upper()
        name = clean(cell(row, 1))
        if not code or not name:
            continue
        email_templates.append(
            {
                "projectCode": code,
                "name": name,
                "triggerType": normalize_trigger(cell(row, 2)),
                "offsetHours": int(cell(row, 3) or 0),
                "subject": replace_vars(cell(row, 4)),
                "body": replace_vars(cell(row, 5)),
                "status": clean(cell(row, 6)) or "啟用",
                "senderName": clean(cell(row, 7)) or "DM Chen",
                "stopWhenBooked": True,
            }
        )

    leads = []
    if "光哥_測驗名單" in sheets:
        for row in nonempty_rows(sheets["光哥_測驗名單"])[1:]:
            email = clean(cell(row, 3)).lower()
            if not email:
                continue
            booked_text = clean(cell(row, 6))
            booked = "已預約" in booked_text and "未預約" not in booked_text
            leads.append(
                {
                    "createdAt": iso_dt(cell(row, 0)),
                    "projectCode": "FIRSTHOME",
                    "projectName": "園區工程師首購計畫",
                    "clientName": clean(cell(row, 2)),
                    "email": email,
                    "phone": clean(cell(row, 4)),
                    "answers": clean(cell(row, 5)),
                    "status": "booked" if booked else "pending",
                }
            )

    payload = {
        "tenantName": "DM Chen",
        "ownerName": "DMtest",
        "ownerEmail": "ncs1491311@gmail.com",
        "projects": projects,
        "questions": questions,
        "consultants": consultants,
        "availability": availability,
        "appointments": appointments,
        "rejected": rejected,
        "waitlist": waitlist,
        "emailTemplates": email_templates,
        "leads": leads,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({key: len(value) for key, value in payload.items() if isinstance(value, list)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
